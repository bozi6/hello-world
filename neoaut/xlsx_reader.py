#  xlsx_reader.py Copyright (C) 2025  Konta Boáz
#      This program comes with ABSOLUTELY NO WARRANTY; for details type `show w'.
#      This is free software, and you are welcome to redistribute it
#      under certain conditions; type `show c' for details.
#   Last Modified: 2025. 06. 20. 22:57

"""
Excel fájl olvasási funkciók - több munkalap támogatással és szűréssel
"""

import openpyxl
from typing import List, Dict, Any
import logging


def read_xlsx_file(file_path: str, filter_null_columns: List[str] = None) -> Dict[str, Dict[str, Any]]:
    """
    Excel fájl beolvasása - minden munkalapon végigmegy és szűri a NULL értékeket
    
    Args:
        file_path (str): Az Excel fájl útvonala
        filter_null_columns (List[str]): Azok az oszlopok, ahol NULL értékek esetén kiszűrjük a sort
        
    Returns:
        Dict: Az Excel adatok munkalapok szerint szervezve
        {
            'sheet_name1': {
                'headers': [],
                'rows': [],
                'column_types': {}
            },
            'sheet_name2': {...}
        }
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        all_sheets_data = {}

        logging.info(f"Talált munkalapok: {workbook.sheetnames}")

        # Végigmegyünk az összes munkalapon
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            logging.info(f"Munkalap feldolgozása: {sheet_name}")

            # Ellenőrizzük, hogy van-e adat a munkalapon
            if sheet.max_row < 2:  # Nincs adat, csak fejléc vagy az sem
                logging.warning(f"A '{sheet_name}' munkalap üres vagy csak fejlécet tartalmaz")
                continue

            # Fejlécek olvasása (első sor)
            headers = []
            for cell in sheet[1]:
                if cell.value is not None:
                    headers.append(str(cell.value).strip())
                else:
                    break

            if not headers:
                logging.warning(f"Nincsenek fejlécek a '{sheet_name}' munkalapon")
                continue

            # Adatok olvasása
            rows = []
            filtered_rows_count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row[:len(headers)]):
                    # Csak a fejlécekkel megegyező számú oszlopot vesszük
                    row_data = []
                    for i, cell in enumerate(row[:len(headers)]):
                        row_data.append(cell if cell is not None else '')

                    # Szűrés NULL értékekre a megadott oszlopokban
                    if filter_null_columns and should_filter_row(row_data, headers, filter_null_columns):
                        filtered_rows_count += 1
                        continue

                    rows.append(row_data)

            if not rows:
                logging.warning(f"Nincsenek adatok a '{sheet_name}' munkalapon (szűrés után)")
                continue

            # Oszlop típusok meghatározása
            column_types = detect_column_types(rows, headers)

            # Munkalap adatok tárolása
            all_sheets_data[sheet_name] = {
                'headers'     : headers,
                'rows'        : rows,
                'column_types': column_types
            }

            logging.info(f"'{sheet_name}' munkalap: {len(rows)} sor, {len(headers)} oszlop "
                         f"(kiszűrt sorok: {filtered_rows_count})")

        workbook.close()

        if not all_sheets_data:
            raise ValueError("Nincsenek feldolgozható adatok az Excel fájlban")

        return all_sheets_data

    except FileNotFoundError:
        logging.error(f"A fájl nem található: {file_path}")
        raise
    except Exception as e:
        logging.error(f"Hiba az Excel fájl olvasása során: {str(e)}")
        raise


def should_filter_row(row_data: List[Any], headers: List[str], filter_columns: List[str]) -> bool:
    """
    Meghatározza, hogy kiszűrjük-e a sort a megadott oszlopokban lévő értékek miatt
    Csak akkor szűr, ha MINDEN szűrendő oszlop üres. Ha bármelyikben van érték, megtartja a sort.
    
    Args:
        row_data (List[Any]): A sor adatai
        headers (List[str]): Az oszlop fejlécek
        filter_columns (List[str]): A szűrendő oszlopok nevei
        
    Returns:
        bool: True, ha ki kell szűrni a sort (minden szűrendő oszlop üres)
    """
    # Ellenőrizzük minden szűrendő oszlopot
    for col_name in filter_columns:
        if col_name in headers:
            col_index = headers.index(col_name)
            if col_index < len(row_data):
                value = row_data[col_index]
                # Ha van érték ebben az oszlopban (nem NULL, nem üres, nem csak whitespace)
                if value is not None and str(value).strip() != '' and str(value).upper() != 'NULL':
                    logging.debug(f"Sor megtartva: '{col_name}' oszlopban van érték: '{value}'")
                    return False  # NEM szűrjük ki, mert van érték
    
    # Ha minden szűrendő oszlop üres volt, akkor kiszűrjük
    logging.debug("Sor kiszűrve: minden szűrendő oszlop üres")
    return True


def read_single_sheet(file_path: str, sheet_name: str = None) -> Dict[str, Any]:
    """
    Egyetlen munkalap beolvasása
    
    Args:
        file_path (str): Az Excel fájl útvonala
        sheet_name (str): A munkalap neve (ha None, akkor az aktív lap)
        
    Returns:
        Dict: Az Excel adatok (headers, rows, column_types)
    """
    try:
        workbook = openpyxl.load_workbook(file_path)

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Nincs '{sheet_name}' nevű munkalap")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
            sheet_name = sheet.title

        # Fejlécek olvasása (első sor)
        headers = []
        for cell in sheet[1]:
            if cell.value is not None:
                headers.append(str(cell.value).strip())
            else:
                break

        if not headers:
            raise ValueError(f"Nincsenek fejlécek a '{sheet_name}' munkalapon")

        # Adatok olvasása
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row[:len(headers)]):
                # Csak a fejlécekkel megegyező számú oszlopot vesszük
                row_data = []
                for i, cell in enumerate(row[:len(headers)]):
                    row_data.append(cell if cell is not None else '')
                rows.append(row_data)

        # Oszlop típusok meghatározása
        column_types = detect_column_types(rows, headers)

        workbook.close()

        return {
            'headers'     : headers,
            'rows'        : rows,
            'column_types': column_types
        }

    except FileNotFoundError:
        logging.error(f"A fájl nem található: {file_path}")
        raise
    except Exception as e:
        logging.error(f"Hiba az Excel fájl olvasása során: {str(e)}")
        raise


def detect_column_types(rows: List[List[Any]], headers: List[str]) -> Dict[str, str]:
    """
    Oszlop típusok automatikus felismerése az adatok alapján
    
    Args:
        rows (List[List[Any]]): Az adatsorok
        headers (List[str]): Az oszlop fejlécek
        
    Returns:
        Dict[str, str]: Oszlop típusok
    """
    column_types = {}

    for col_idx, header in enumerate(headers):
        # Első néhány nem üres érték vizsgálata a típus meghatározásához
        sample_values = []
        for row in rows[:min(100, len(rows))]:  # Maximum 100 minta
            if col_idx < len(row) and row[col_idx] != '':
                sample_values.append(row[col_idx])

        if not sample_values:
            column_types[header] = 'string'
            continue

        # Típus meghatározása
        is_int = all(isinstance(val, int) for val in sample_values)
        is_float = all(isinstance(val, (int, float)) for val in sample_values)
        is_date = any('date' in str(type(val)).lower() for val in sample_values)

        if is_date:
            column_types[header] = 'datetime'
        elif is_int:
            column_types[header] = 'int'
        elif is_float:
            column_types[header] = 'float'
        else:
            column_types[header] = 'string'

    return column_types


def sanitize_sheet_name_for_table(sheet_name: str) -> str:
    """
    Munkalap név tisztítása MySQL tábla névhez
    
    Args:
        sheet_name (str): A munkalap eredeti neve
        
    Returns:
        str: Tisztított tábla név
    """
    import re
    # Speciális karakterek eltávolítása/cseréje
    clean_name = re.sub(r'[^\w]', '_', sheet_name.strip())
    clean_name = re.sub(r'_+', '_', clean_name)  # Többszörös _ csökkentése
    clean_name = clean_name.strip('_')  # Kezdő/záró _ eltávolítása

    # MySQL reserved words elkerülése
    reserved_words = ['select', 'from', 'where', 'insert', 'update', 'delete', 'table', 'al']
    if clean_name.lower() in reserved_words:
        clean_name = f"{clean_name}_sheet"

    # Üres név kezelése
    if not clean_name:
        clean_name = "sheet_data"

    return clean_name.lower()