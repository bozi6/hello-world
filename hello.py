from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import math

wb = Workbook()
# destination filename
dest_filename = 'sample.xlsx'

# select active worksheet (ws1)
ws1 = wb.active
ws1.title = "range names"
# Create numbers in a range in worksheet1
for row in range(1, 35):
    ws1.append(range(28))
# Create another worksheet and name it Pi
ws2 = wb.create_sheet(title="Pi")

# Enter in b2 and f5 cells the pi
ws2['B2'] = 3.14
ws2['F5'] = math.pi

# Create another worksheet and name it Data
ws3 = wb.create_sheet(title="Data")
for row in range(4, 20):
    for col in range(3, 10):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
# Print E11 cell value in worksheet 3
print(ws3['E11'].value)

# Save the excel file to current directory
wb.save(filename=dest_filename)
