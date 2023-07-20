import logging  # Loggoláshoz modul
from termcolor import colored
import re  # Reguláris kifejezésekhez modul
from xml.dom import minidom  # xml fájl olvasásához modul

from currency_converter import CurrencyConverter # Valutaváltó
from openpyxl import Workbook  # Excel táblázat kezeléshez modul
from openpyxl.formatting.rule import Rule
from openpyxl.styles import NamedStyle, Font, PatternFill  # Stílus importálása modul
from openpyxl.styles.differential import DifferentialStyle

FORMAT = '%(levelname)s: %(asctime)-2s %(message)s'  # Loggolás formátumának beállításra
logging.basicConfig(format=FORMAT, level=logging.INFO)  # loggolás beálltása INFO-ra(csak a lényeg)

olvasando_fajl = "sms_full.xml"
kiirando_fajl = "./balance.xlsx"

# xml fájl betöltése
mydoc = minidom.parse(olvasando_fajl)

c = CurrencyConverter()

# Munkalap stílusának beálítása forintra
still = NamedStyle(name="Pénzecske")  # Ilyen nevű stílus hozzáadása
still.number_format = '#,##0 "HUF";-#,##0 "HUF"'  # A számformátum beállítása forintra
diff_style = DifferentialStyle(fill=PatternFill(bgColor='C6EFCE', fgColor='006100'))  # A feltételes formázás beállítása
rule = Rule(type="expression", dxf=diff_style)  # Feltételes kifejezés megadása
rule.formula = ["$B2>0"]  # Formula a felt. formázáshoz


def penzvalto(mit):
    """ Árfolyam váltó függvény
    a bejövő mit hez 123,45 GBP
    kiszámolja, hogy az mennyi forintban """
    valuta = mit[-3:]  # A bejövő valuta megnevezésének eltávolítása
    mennyi = mit[:-3].strip().replace(',', '.')
    ennyi = c.convert(mennyi, valuta, "HUF")
    return round(ennyi, 2)


def ujlap_letrehozasa(lap_elnevezese):
    """ Új munkalap létrehozása a munkafüzetben. """
    logging.debug('ujlap funkció meghívva')
    """Alapértékek megadása"""
    szotar = {'A1': 'Dátum', 'A2': 'Nyitó összeg', 'B1': 'Pénzmozgás',
              'C1': 'Valuta', 'D1': 'LU egyenleg', 'E1': 'Üzenet',
              'F1': 'Egyenleg', 'G1': 'Ez hibádzik'}
    for kulcs, ertek in szotar.items():
        ws[kulcs] = ertek
        ws[kulcs].font = Font(bold=True)
    ws.title = lap_elnevezese  # Az új lap elnevezése
    ws['B2'].font = Font(italic=True, bold=True)
    szelessegek = {'A': 25, 'B': 15, 'C': 15, 'D': 15, 'E': 160, 'F': 13, 'G': 13}
    for kulcs, ertek in szelessegek.items():
        ws.column_dimensions[kulcs].width = ertek
    ws.cell(row=2, column=2, value='=D3-B3').style = still
    # A B2 cellába beírjuk a kezdőértéket első sms egyenlegéből kivonva az első levonást = nyitó egyenleg
    ws.sheet_view.zoomScale = 170  # A nézet kinagyítása az aktuális oldalon


# Munkalap létrehozása a memóriában
wb = Workbook()
wb.add_named_style(still)
ws = wb.active

items = mydoc.getElementsByTagName('sms')  # smsek beolvasása az items-be

lapnev = items[0].attributes['readable_date'].value  # Az első sms dátumának kinyerése
lapnev = lapnev[:4]  # Csak az évszám kivágása

cserek = [
    ("Egy.:", "Egyenleg:"),
    ("Egy:", "Egyenleg:"),
    ("Egy:+", "Egyenleg:"),
    ("Egyenleg:+", "Egyenleg: "),
    (",-HUF", " HUF"),
]

ujlap_letrehozasa(lapnev)  # beküldjük az új laphoz
sor = 3  # a táblázat írni kívánt első sora
i = 1  # belső változó
zd = 0 # debug változó
for elem in items:  # SMS-ek beolvasása
    # if 'SIKERTELEN' in elem.attributes['body'].value:  # Ha a SIKERTELEN üzenetet kaptuk, akkor nem törődünk vele
    #     continue
    tel = elem.attributes['address'].value  # telefonszám kinyerése az üzenetből, később lehet szűrni.
    rd = elem.attributes['readable_date'].value  # Dátum hozzáadása az rd változóhoz
    bd = elem.attributes['body'].value  # Az üzenet szövege
    cn = elem.attributes['contact_name'].value  # Az üzenetküldő neve
    penz_pattern = r'(-|\+)[0-9]+(,?|\.?)[0-9]+\s?,?-?[A-Z][A-Z][A-Z]'
    egyenleg_pattern = r'(Egyenleg:\s)(\+?[0-9]+.[0-9]+\.?[0-9]+)'
    # előjel - vagy +, utána 0 vagy több szám, aztán 0 vagy egy karakter,
    # 0 vagy több szám, 0 vagy 1 szóköz 0 vagy 1 , 0 vagy 1 - és három karakter a pénznem azonosításhoz.
    # x = re.findall(pattern, bd)
    aktev = rd[:4]  # Az aktális évszám kinyerése a dátumból.
    if aktev != ws.title:  # Ha ez nem egyezik a lap nevével akkor új lapot nyitunk
        ws.conditional_formatting.add("B2:B{}".format(ws.max_row), rule)
        # összeadjuk a bevételeket. Ezt még az aktuális és nem új lapon tesszük.
        ws.cell(row=ws.max_row + 1, column=1, value='Bevétel:')
        ws.cell(row=ws.max_row + 1, column=1, value='=SUMIF(B2:B{},">0")'.format(ws.max_row - 2)).style = still
        ws.cell(row=ws.max_row + 1, column=1, value='Kiadás:')
        ws.cell(row=ws.max_row + 1, column=1, value='=SUMIF(B2:B{},"<0")'.format(ws.max_row - 4)).style = still
        ws = wb.create_sheet()  # itt csinálunk új lapot
        wb.active  # itt pedig aktívvá tesszük.
        ujlap_letrehozasa(aktev)  # meghívjuk rá az alapbeállításokat.
        sor = 3
    try:
        x = re.search(penz_pattern, bd).group()  # megkeressük a pénzeket az üzenetből
        for old, new in cserek:
            bd = bd.replace(old, new)
        egyenleg = re.search(egyenleg_pattern, bd).group(2)
        logging.debug('Egyenleg értéke: {} - Üzi: {}'.format(egyenleg, bd))
        x = x.replace(".", "")  # kivesszük a pontot (ezres elválasztó) az összegek közül
        x = x.replace(",-HUF", " HUF")  # Ha az üzenetben ,-HUF van azt átalakítjuk HUf-ra
        if x.find("HUF") > -1:  # Ha benne van az összegben a HUF => forint alapú a dolog
            x = x.replace(" HUF", "")  # Kivesszük a HUF szöveget az excel pénznem felismerése miatt.
            ws.cell(row=sor, column=2, value=float(x)).style = still  # beírjuk az aktuális sorba a második oszlopba
        else:  # ha nem HUF-ban van megadva a pénz
            ws.cell(row=sor, column=3, value=x)  # Beírjuk a harmadik oszlopba az összeget
            forintban = penzvalto(x)  # Átváltjuk az összeget forintra
            ws.cell(row=sor, column=2, value=forintban).style = still  # Ezt beírjuk a második oszlopba
            logging.debug('Valuta értéke:', ws.cell(row=sor, column=2).value)
        egyenleg = egyenleg.replace(".", "")
        ws.cell(row=sor, column=1, value=rd)  # A dátumot beírjuk az első oszlopba
        bd = bd.replace("...", "")  # Az üzenet elejéről a pontokat kivesszük még.
        logging.debug(bd)
        ws.cell(row=sor, column=4, value=float(egyenleg)).style = still  # Az egyenleget beírjuk a negyedik oszlopba
        ws.cell(row=sor, column=5, value=bd)  # Az üzenetet beírjuk az ötödik oszlopba, ellenőrzés czéljából
        ws.cell(row=sor, column=6, value="=SUM(D{},B{})".format(sor, sor + 1)).style = still
        # Nem mindegy, hogy a képletbe elválsztónak , vagy ; van itt a ; hibát dob az excelben.
        ws.cell(row=sor, column=7, value="=D{}-F{}".format(sor + 1, sor)).style = still
        logging.debug("Sorszám:{}; Dátum: {} - Érték: {} - Üzi: {}".format(i, rd, x, bd))
        # ha DEBUG-ra van állítva a loggolás
        i += 1  # sorszám növelése
    except AttributeError:
        logging.info("Sorszám:{}; - {} - Nem átutalásos sms - {}".format(i, rd, bd))
        continue
    sor += 1
    # Az összegeket az oszlop végén öszeadjuk és beírjuk az eredményt. és hozzáadjuk a stílust
    ws.cell(row=ws.max_row + 1, column=2, value="=SUM(B2:B{})".format(ws.max_row)).style = still # B oszlop összegzése
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True, italic=True)  # A betűtípust vastaggá tesszük az eredményen
    # logging.debug('Munkalap max sorszáma: {}'.format(ws.max_row))
ws.conditional_formatting.add("B2:B{}".format(ws.max_row), rule) # feltételes formázás hozzáadása
# összeadjuk a bevételeket.
ws.cell(row=ws.max_row + 1, column=1, value='Bevétel:')
ws.cell(row=ws.max_row + 1, column=1, value='=SUMIF(B2:B{},">0")'.format(ws.max_row - 2)).style = still
ws.cell(row=ws.max_row + 1, column=1, value='Kiadás:')
ws.cell(row=ws.max_row + 1, column=1, value='=SUMIF(B2:B{},"<0")'.format(ws.max_row - 4)).style = still
wb.save(kiirando_fajl)  # táblázat elmentése
print('Elkészült munkalapok:')
for sheet in wb:
    print(sheet.title)
