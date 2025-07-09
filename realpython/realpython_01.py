"""Táblázat olvasgató."""
from datetime import datetime
from openpyxl import load_workbook
from classes import Product, Review, Strange
from mapping import (
    PRODUCT_ID,
    PRODUCT_PARENT,
    PRODUCT_TITLE,
    PRODUCT_CATEGORY,
    REVIEW_DATE,
    REVIEW_ID,
    REVIEW_CUSTOMER,
    REVIEW_STARS,
    REVIEW_HEADLINE,
    REVIEW_BODY,
    REVIEW_DATE,
    HELPFUL_VOTES,
    TOTAL_VOTES,
    VINE,
    VERIFIED_PURCHASE
)


def parse_date(raw_date):
    """Dátum parse-olása stringből datetime formátumra."""
    if isinstance(raw_date, datetime):
        return raw_date
    try:
        return datetime.strptime(raw_date, "%Y-%m-%d")  # Helyesen használva a datetime.strptime metódus.
    except ValueError:
        return None


def create_product(row):
    """Új Product objektum létrehozása egy adott sor alapján."""
    return Product(
        product_id=row[PRODUCT_ID],
        parent=row[PRODUCT_PARENT],
        title=row[PRODUCT_TITLE],
        product_category=row[PRODUCT_CATEGORY],
        date=None,  # vagy adj meg alapértelmezett értéket
    )


def create_review(row):
    """Új Review objektum létrehozása egy adott sor alapján."""
    return Review(
        id=row[REVIEW_ID],
        customer_id=row[REVIEW_CUSTOMER],
        stars=row[REVIEW_STARS],
        headline=row[REVIEW_HEADLINE],
        body=row[REVIEW_BODY],
        date=parse_date(row[REVIEW_DATE]),
    )


def create_strange(row):
    """Új strange objektum létrehozása egy adott sort alapján"""
    return Strange(
        helpvotes=row[HELPFUL_VOTES],
        tot_votes=row[TOTAL_VOTES],
        vine=row[VINE],
        veripurch=row[VERIFIED_PURCHASE]
    )


if __name__ == "__main__":
    wb = load_workbook(filename="reviews-sample.xlsx", read_only=True)
    sh = wb.active

    products = []
    reviews = []
    strange = []
    for row in sh.iter_rows(min_row=2, values_only=True):
        products.append(create_product(row))
        reviews.append(create_review(row))
        strange.append(create_strange(row))

    print(products[0].title)
    print(reviews[0].headline)
    print(strange[2])


