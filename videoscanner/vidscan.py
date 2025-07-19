#  vidscan.py Copyright (C) 2025  Konta Boáz
#      This program comes with ABSOLUTELY NO WARRANTY; for details type `show w'.
#      This is free software, and you are welcome to redistribute it
#      under certain conditions; type `show c' for details.
#   Last Modified: 2025. 06. 19. 18:40
# Scans a direcotry for video files and put it into an Excel file for easy categorizing

import argparse
import json
import subprocess
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill


class VideoScanner:
    def __init__(self):
        # Videó fájlkiterjesztések
        self.video_extensions = {
            '.mp4', '.avi', '.mkv', '.mov', '.wmv', '.flv', '.webm', '.m4v',
            '.mpg', '.mpeg', '.3gp', '.ogv', '.ts', '.mts', '.m2ts'
        }

        # Excel munkafüzet és munkalapok
        self.workbook = openpyxl.Workbook()
        self.main_sheet = self.workbook.active
        self.main_sheet.title = "Összes videó"
        self.year_sheets = {}

        # Fájlszámláló
        self.file_counter = 0

        self._setup_main_sheet()

    def _setup_main_sheet(self):
        """Beállítja a fő munkalap fejléceit"""
        headers = ['Sorszám', 'Fájl elérési út', 'Fájlnév', 'Hossz (óó:pp:mm)', 'Felbontás', 'Módosítás dátuma', 'Fájlméret (MB)']

        for col, header in enumerate(headers, 1):
            cell = self.main_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Oszlopszélességek beállítása
        self.main_sheet.column_dimensions['A'].width = 10
        self.main_sheet.column_dimensions['B'].width = 60
        self.main_sheet.column_dimensions['C'].width = 40
        self.main_sheet.column_dimensions['D'].width = 15
        self.main_sheet.column_dimensions['E'].width = 15
        self.main_sheet.column_dimensions['F'].width = 20
        self.main_sheet.column_dimensions['G'].width = 15

    def _setup_year_sheet(self, year):
        """Beállít egy év munkalapot"""
        if year not in self.year_sheets:
            sheet = self.workbook.create_sheet(title=str(year))
            self.year_sheets[year] = sheet

            headers = ['Sorszám', 'Fájl elérési út', 'Fájlnév', 'Hossz (óó:pp:mm)', 'Felbontás', 'Módosítás dátuma',
                       'Fájlméret (MB)']

            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDFF", end_color="DDDDFF", fill_type="solid")

            # Oszlopszélességek beállítása
            sheet.column_dimensions['A'].width = 10
            sheet.column_dimensions['B'].width = 60
            sheet.column_dimensions['C'].width = 40
            sheet.column_dimensions['D'].width = 15
            sheet.column_dimensions['E'].width = 15
            sheet.column_dimensions['F'].width = 20
            sheet.column_dimensions['G'].width = 15

    def get_video_info(self, file_path):
        """Meghatározza a videó hosszát és felbontását ffprobe segítségével"""
        try:
            cmd = [
                'ffprobe', '-v', 'quiet', '-print_format', 'json',
                '-show_format', '-show_streams', str(file_path)
            ]

            result = subprocess.run(cmd, capture_output=True, text=True)

            if result.returncode == 0:
                data = json.loads(result.stdout)
                
                # Videó hossz meghatározása
                duration = float(data['format']['duration'])
                hours = int(duration // 3600)
                minutes = int((duration % 3600) // 60)
                seconds = int(duration % 60)
                duration_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                
                # Videó felbontás meghatározása
                resolution = "Ismeretlen"
                for stream in data.get('streams', []):
                    if stream.get('codec_type') == 'video':
                        width = stream.get('width')
                        height = stream.get('height')
                        if width and height:
                            resolution = f"{width}x{height}"
                            break

                return {
                    'duration': duration_str,
                    'resolution': resolution
                }
            else:
                return {
                    'duration': "Ismeretlen",
                    'resolution': "Ismeretlen"
                }
        except (subprocess.SubprocessError, json.JSONDecodeError, KeyError, ValueError):
            return {
                'duration': "Hiba",
                'resolution': "Hiba"
            }

    def get_file_info(self, file_path):
        """Lekéri a fájl információit"""
        try:
            stat = file_path.stat()
            modification_time = datetime.fromtimestamp(stat.st_mtime)
            size_mb = round(stat.st_size / (1024 * 1024), 2)

            return {
                'modification_date': modification_time.strftime('%Y-%m-%d %H:%M:%S'),
                'modification_year': modification_time.year,
                'size_mb'          : size_mb
            }
        except OSError:
            return {
                'modification_date': 'Ismeretlen',
                'modification_year': None,
                'size_mb'          : 0
            }

    def add_video_to_sheets(self, file_path, file_info, video_info):
        """Hozzáadja a videót a megfelelő munkalapokhoz"""
        self.file_counter += 1

        row_data = [
            self.file_counter,
            str(file_path.parent),
            file_path.name,
            video_info['duration'],
            video_info['resolution'],
            file_info['modification_date'],
            file_info['size_mb']
        ]

        # Hozzáadás a fő munkalaphoz
        row = self.main_sheet.max_row + 1
        for col, value in enumerate(row_data, 1):
            self.main_sheet.cell(row=row, column=col, value=value)

        # Hozzáadás az év munkalaphoz
        if file_info['modification_year']:
            year = file_info['modification_year']
            self._setup_year_sheet(year)

            year_sheet = self.year_sheets[year]
            row = year_sheet.max_row + 1
            for col, value in enumerate(row_data, 1):
                year_sheet.cell(row=row, column=col, value=value)

    def scan_directory(self, directory_path):
        """Átvizsgál egy könyvtárat videófájlok után"""
        directory = Path(directory_path)

        if not directory.exists():
            print(f"Hiba: A könyvtár nem létezik: {directory_path}")
            return

        print(f"Könyvtár vizsgálata: {directory_path}")

        try:
            for file_path in directory.rglob('*'):
                if file_path.is_file() and file_path.suffix.lower() in self.video_extensions:
                    print(f"Videófájl feldolgozása: {file_path.name}")

                    # Fájl információk lekérése
                    file_info = self.get_file_info(file_path)

                    # Videó információk lekérése (hossz és felbontás)
                    video_info = self.get_video_info(file_path)

                    # Hozzáadás a munkalapokhoz
                    self.add_video_to_sheets(file_path, file_info, video_info)

        except PermissionError as e:
            print(f"Engedély hiba: {e}")
        except Exception as e:
            print(f"Hiba történt: {e}")

    def sort_excel_sheets_by_year(self):
        """Rendezés a munkalapok között év szerint növekvő sorrendbe."""
        # Gyűjtse össze az év munkalapokat és sorrendezze a lap neve szerint.
        year_sheets_sorted = sorted(
            self.year_sheets.items(),
            key=lambda item: int(item[0])  # Év számként rendezése.
        )

        # Rendezett sorrendben helyezze át a lapokat a Workbook-ban.
        for index, (year, sheet) in enumerate(year_sheets_sorted, start=1):
            self.workbook.move_sheet(sheet, offset=index - len(self.workbook.sheetnames))
    
    def save_excel(self, output_file="video_files.xlsx"):
        """Elmenti az Excel fájlt, rendezve az év munkalapokat."""
        try:
            # Munkalapok közötti rendezés.
            self.sort_excel_sheets_by_year()

            # Ha nincsenek évek munkalapjai, töröljük az alapértelmezett üres lapot.
            if not self.year_sheets and len(self.workbook.worksheets) > 1:
                self.workbook.remove(self.workbook.worksheets[1])

            self.workbook.save(output_file)
            print(f"Excel fájl mentve: {output_file}")
            print(f"Összesen {self.file_counter} videófájl feldolgozva")
            print(f"Létrehozott munkalapok: {len(self.workbook.worksheets)}")

        except Exception as e:
            print(f"Hiba az Excel fájl mentésekor: {e}")


def main():
    parser = argparse.ArgumentParser(description='Videófájlok keresése és Excel táblázat készítése')
    parser.add_argument('directories', nargs='+', help='Vizsgálandó könyvtárak')
    parser.add_argument('-o', '--output', default='video_files.xlsx', help='Kimeneti Excel fájl neve')

    args = parser.parse_args()

    # Ellenőrizzük, hogy az ffprobe elérhető-e
    try:
        subprocess.run(['ffprobe', '-version'], capture_output=True, check=True)
    except (subprocess.SubprocessError, FileNotFoundError):
        print("Figyelem: ffprobe nem található. A videók hossza és felbontása 'Hiba'-ként jelenik meg.")
        print("FFmpeg telepítése szükséges a videó információk meghatározásához.")

    scanner = VideoScanner()

    for directory in args.directories:
        scanner.scan_directory(directory)

    scanner.save_excel(args.output)


if __name__ == "__main__":
    main()